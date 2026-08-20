from __future__ import annotations

import abc
import csv
import hashlib
import json
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook


PROVENANCE_OFFICIAL_GOVERNMENT = "official_government"
PROVENANCE_OFFICIAL_SST = "official_sst"
PROVENANCE_SYNTHETIC_TEST = "synthetic_test"
PROVENANCE_MANUAL_UNVERIFIED = "manual_unverified"


@dataclass
class NormalizedTaxRecord:
    state: str
    jurisdiction_type: str
    jurisdiction_code: str
    jurisdiction_name: str
    tax_code: str
    general_rate_bps: int
    grocery_rate_bps: int
    prepared_rate_bps: int
    confidence: str
    effective_from: date
    effective_to: date | None
    assignment_key_type: str
    assignment_key: str
    assignment_precision: str


@dataclass
class NormalizedTaxDataset:
    source_key: str
    source_type: str
    source_name: str
    source_reference: str
    version_tag: str
    effective_from: date
    effective_to: date | None
    records: list[NormalizedTaxRecord]


class TaxDataAdapter(abc.ABC):
    source_key: str

    @abc.abstractmethod
    def discover_version(self, source_path: Path) -> str:
        raise NotImplementedError

    @abc.abstractmethod
    def parse_source(self, source_path: Path) -> NormalizedTaxDataset:
        raise NotImplementedError

    @abc.abstractmethod
    def validate_records(self, dataset: NormalizedTaxDataset) -> list[str]:
        raise NotImplementedError


class PublicStateRatesAdapter(TaxDataAdapter):
    source_key = "public_state_rates"

    def discover_version(self, source_path: Path) -> str:
        payload = json.loads(source_path.read_text(encoding="utf-8"))
        return str(payload.get("version_tag") or "unknown")

    def parse_source(self, source_path: Path) -> NormalizedTaxDataset:
        payload = json.loads(source_path.read_text(encoding="utf-8"))
        effective_from = date.fromisoformat(str(payload["effective_from"]))
        effective_to = date.fromisoformat(str(payload["effective_to"])) if payload.get("effective_to") else None

        records: list[NormalizedTaxRecord] = []
        for row in payload.get("states") or []:
            state = str(row.get("state") or "").strip().upper()
            records.append(
                NormalizedTaxRecord(
                    state=state,
                    jurisdiction_type="state",
                    jurisdiction_code=f"STATE:{state}",
                    jurisdiction_name=state,
                    tax_code=f"{state}-STATE",
                    general_rate_bps=int(row.get("general_rate_bps") or 0),
                    grocery_rate_bps=int(row.get("grocery_rate_bps") or 0),
                    prepared_rate_bps=int(row.get("prepared_rate_bps") or int(row.get("general_rate_bps") or 0)),
                    confidence=str(row.get("confidence") or "medium"),
                    effective_from=effective_from,
                    effective_to=effective_to,
                    assignment_key_type="state",
                    assignment_key=state,
                    assignment_precision="STATE_ONLY",
                )
            )

        return NormalizedTaxDataset(
            source_key=str(payload.get("source_key") or self.source_key),
            source_type=str(payload.get("source_type") or PROVENANCE_MANUAL_UNVERIFIED),
            source_name=str(payload.get("source_name") or "Public state-level rates"),
            source_reference=str(payload.get("source_reference") or ""),
            version_tag=str(payload.get("version_tag") or "unknown"),
            effective_from=effective_from,
            effective_to=effective_to,
            records=records,
        )

    def validate_records(self, dataset: NormalizedTaxDataset) -> list[str]:
        errors: list[str] = []
        seen: set[tuple[str, str, date]] = set()
        for record in dataset.records:
            if len(record.state) != 2:
                errors.append(f"invalid state: {record.state}")
            if record.general_rate_bps < 0 or record.grocery_rate_bps < 0 or record.prepared_rate_bps < 0:
                errors.append(f"negative rate in {record.state}")
            if record.general_rate_bps > 2000 or record.grocery_rate_bps > 2000 or record.prepared_rate_bps > 2000:
                errors.append(f"rate too high in {record.state}")
            key = (record.jurisdiction_code, record.tax_code, record.effective_from)
            if key in seen:
                errors.append(f"duplicate contradictory record: {key}")
            seen.add(key)
        return errors


class SstMemberStateAdapter(TaxDataAdapter):
    source_key = "sst_member_sample"

    def discover_version(self, source_path: Path) -> str:
        payload = json.loads(source_path.read_text(encoding="utf-8"))
        return str(payload.get("version_tag") or "unknown")

    def parse_source(self, source_path: Path) -> NormalizedTaxDataset:
        payload = json.loads(source_path.read_text(encoding="utf-8"))
        effective_from = date.fromisoformat(str(payload["effective_from"]))
        effective_to = date.fromisoformat(str(payload["effective_to"])) if payload.get("effective_to") else None

        records: list[NormalizedTaxRecord] = []
        for row in payload.get("records") or []:
            assignment = row.get("assignment") or {}
            records.append(
                NormalizedTaxRecord(
                    state=str(row.get("state") or "").strip().upper(),
                    jurisdiction_type=str(row.get("jurisdiction_type") or "state").strip(),
                    jurisdiction_code=str(row.get("jurisdiction_code") or "").strip(),
                    jurisdiction_name=str(row.get("jurisdiction_name") or "").strip(),
                    tax_code=str(row.get("tax_code") or "").strip(),
                    general_rate_bps=int(row.get("general_rate_bps") or 0),
                    grocery_rate_bps=int(row.get("grocery_rate_bps") or 0),
                    prepared_rate_bps=int(row.get("prepared_rate_bps") or int(row.get("general_rate_bps") or 0)),
                    confidence="synthetic",
                    effective_from=effective_from,
                    effective_to=effective_to,
                    assignment_key_type=str(assignment.get("key_type") or "zip5").strip(),
                    assignment_key=str(assignment.get("key") or "").strip(),
                    assignment_precision=str(assignment.get("precision") or "ZIP5").strip(),
                )
            )

        return NormalizedTaxDataset(
            source_key=str(payload.get("source_key") or self.source_key),
            source_type=str(payload.get("source_type") or PROVENANCE_SYNTHETIC_TEST),
            source_name=str(payload.get("source_name") or "SST member-state sample"),
            source_reference=str(payload.get("source_reference") or ""),
            version_tag=str(payload.get("version_tag") or "unknown"),
            effective_from=effective_from,
            effective_to=effective_to,
            records=records,
        )

    def validate_records(self, dataset: NormalizedTaxDataset) -> list[str]:
        errors: list[str] = []
        for record in dataset.records:
            if not record.tax_code:
                errors.append("missing tax code")
            if not record.jurisdiction_code:
                errors.append("missing jurisdiction code")
            if record.assignment_key_type == "zip5" and len(record.assignment_key) != 5:
                errors.append(f"malformed ZIP assignment: {record.assignment_key}")
        return errors


class MissouriDorQ3Adapter(TaxDataAdapter):
    source_key = "mo_dor_q3_2026"

    _STATE_FIPS = 29
    _TARGET_ZIPS = {"65084", "65026"}
    _TARGET_CITY_CODES = {
        "21484": "ELDON",
        "75922": "VERSAILLES",
    }
    _SOURCE_REFERENCE = (
        "https://dor.mo.gov/taxation/business/tax-types/sales-use/documents/dor-rate-database-q3-2026.xlsx;"
        "https://dor.mo.gov/taxation/business/tax-types/sales-use/documents/dor-jurisdiction-total-rate-q3-2026.xlsx;"
        "https://dor.mo.gov/taxation/business/tax-types/sales-use/documents/dor-boundary-database-q3-2026.csv;"
        "https://dor.mo.gov/taxation/business/tax-types/sales-use/documents/Q3-2026-Tax-Breakdown-Report.xlsx"
    )

    def discover_version(self, source_path: Path) -> str:
        directory = self._coerce_directory(source_path)
        for child in directory.iterdir():
            name = child.name.lower()
            if "q3-2026" in name or "q3_2026" in name:
                return "2026Q3"
        return "unknown"

    def parse_source(self, source_path: Path) -> NormalizedTaxDataset:
        directory = self._coerce_directory(source_path)
        rate_path = directory / "dor-rate-database-q3-2026.xlsx"
        jurisdiction_path = directory / "dor-jurisdiction-total-rate-q3-2026.xlsx"
        boundary_path = directory / "dor-boundary-database-q3-2026.csv"
        breakdown_path = directory / "Q3-2026-Tax-Breakdown-Report.xlsx"

        if not rate_path.exists() or not jurisdiction_path.exists() or not boundary_path.exists() or not breakdown_path.exists():
            raise ValueError("missing required Missouri DOR Q3 2026 source files")

        state_general_bps, state_food_bps = self._load_state_rate_bps(rate_path)
        jurisdiction_rates = self._load_jurisdiction_rates(jurisdiction_path)
        city_rates = self._load_city_rate_rollups(jurisdiction_rates)
        zip_rates = self._load_target_zip_rollups(boundary_path, jurisdiction_rates)
        effective_from = date(2026, 7, 1)

        records: list[NormalizedTaxRecord] = [
            NormalizedTaxRecord(
                state="MO",
                jurisdiction_type="state",
                jurisdiction_code="STATE:MO",
                jurisdiction_name="Missouri",
                tax_code="MO-STATE",
                general_rate_bps=state_general_bps,
                grocery_rate_bps=state_food_bps,
                prepared_rate_bps=state_general_bps,
                confidence="high",
                effective_from=effective_from,
                effective_to=None,
                assignment_key_type="state",
                assignment_key="MO",
                assignment_precision="STATE_ONLY",
            )
        ]

        for city_code, city_name in self._TARGET_CITY_CODES.items():
            rates = city_rates.get(city_code)
            if rates is None:
                continue
            records.append(
                NormalizedTaxRecord(
                    state="MO",
                    jurisdiction_type="city",
                    jurisdiction_code=f"CITY:MO:{city_name}",
                    jurisdiction_name=f"{city_name}, MO",
                    tax_code=f"MO-CITY-{city_name}",
                    general_rate_bps=rates["general"],
                    grocery_rate_bps=rates["grocery"],
                    prepared_rate_bps=rates["general"],
                    confidence="high",
                    effective_from=effective_from,
                    effective_to=None,
                    assignment_key_type="city_state",
                    assignment_key=f"{city_name.lower()}|MO",
                    assignment_precision="CITY_COUNTY",
                )
            )

        for zip_code in sorted(self._TARGET_ZIPS):
            rates = zip_rates.get(zip_code)
            if rates is None:
                continue
            records.append(
                NormalizedTaxRecord(
                    state="MO",
                    jurisdiction_type="postal",
                    jurisdiction_code=f"ZIP5:MO:{zip_code}",
                    jurisdiction_name=f"ZIP {zip_code}, MO",
                    tax_code=f"MO-ZIP5-{zip_code}",
                    general_rate_bps=rates["general"],
                    grocery_rate_bps=rates["grocery"],
                    prepared_rate_bps=rates["general"],
                    confidence="high",
                    effective_from=effective_from,
                    effective_to=None,
                    assignment_key_type="zip5",
                    assignment_key=zip_code,
                    assignment_precision="ZIP5",
                )
            )

        # Ensure the source workbook exists; it is part of the provenance bundle.
        _ = source_hash(breakdown_path)

        return NormalizedTaxDataset(
            source_key=self.source_key,
            source_type=PROVENANCE_OFFICIAL_GOVERNMENT,
            source_name="Missouri Department of Revenue Q3 2026 official rates",
            source_reference=self._SOURCE_REFERENCE,
            version_tag="2026Q3",
            effective_from=effective_from,
            effective_to=None,
            records=records,
        )

    def validate_records(self, dataset: NormalizedTaxDataset) -> list[str]:
        errors: list[str] = []
        if dataset.source_type != PROVENANCE_OFFICIAL_GOVERNMENT:
            errors.append("missouri adapter must emit official_government provenance")
        if dataset.version_tag != "2026Q3":
            errors.append("missouri adapter must emit 2026Q3 version")
        if "dor.mo.gov" not in dataset.source_reference:
            errors.append("missouri adapter source reference must cite dor.mo.gov")

        seen_keys: set[tuple[str, str]] = set()
        for record in dataset.records:
            if record.state != "MO":
                errors.append(f"unexpected state in record: {record.state}")
            if record.assignment_key_type == "zip5" and len(record.assignment_key) != 5:
                errors.append(f"malformed ZIP assignment: {record.assignment_key}")
            if record.general_rate_bps < 0 or record.grocery_rate_bps < 0:
                errors.append(f"negative rate in {record.tax_code}")
            key = (record.assignment_key_type, record.assignment_key)
            if key in seen_keys:
                errors.append(f"duplicate assignment key: {key}")
            seen_keys.add(key)

        if not any(r.assignment_key_type == "zip5" and r.assignment_key == "65084" for r in dataset.records):
            errors.append("missing ZIP5 assignment for 65084")
        if not any(r.assignment_key_type == "zip5" and r.assignment_key == "65026" for r in dataset.records):
            errors.append("missing ZIP5 assignment for 65026")
        return errors

    def _coerce_directory(self, source_path: Path) -> Path:
        return source_path if source_path.is_dir() else source_path.parent

    def _load_state_rate_bps(self, rate_path: Path) -> tuple[int, int]:
        workbook = load_workbook(rate_path, data_only=True, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(min_row=2, values_only=True)
        for row in rows:
            state = self._to_int(cast(Any, row[0]))
            row_type = str(row[1] or "").strip()
            fips = self._to_int(cast(Any, row[2]))
            if state == self._STATE_FIPS and row_type == "45" and fips == self._STATE_FIPS:
                return self._pct_decimal_to_bps(row[4]), self._pct_decimal_to_bps(row[6])
        raise ValueError("state-level Missouri rate row not found in DOR rate database")

    def _load_jurisdiction_rates(self, jurisdiction_path: Path) -> dict[str, dict[str, Any]]:
        workbook = load_workbook(jurisdiction_path, data_only=True, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        out: dict[str, dict[str, Any]] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            state_fips = self._to_int(cast(Any, row[0]))
            if state_fips != self._STATE_FIPS:
                continue
            jurisdiction_code = str(row[8] or "").strip()
            if not jurisdiction_code:
                continue
            out[jurisdiction_code] = {
                "city_fips": str(row[2] or "").strip().zfill(5),
                "county_fips": str(row[1] or "").strip().zfill(3),
                "general": self._pct_string_to_bps(row[9]),
                "grocery": self._pct_string_to_bps(row[11]),
            }
        if not out:
            raise ValueError("no Missouri jurisdiction rows found in DOR jurisdiction workbook")
        return out

    def _load_city_rate_rollups(self, jurisdiction_rates: dict[str, dict[str, Any]]) -> dict[str, dict[str, int]]:
        rollups: dict[str, dict[str, int]] = {}
        for row in jurisdiction_rates.values():
            city_fips = str(row["city_fips"])
            if city_fips == "00000":
                continue
            current = rollups.setdefault(city_fips, {"general": 0, "grocery": 0})
            current["general"] = max(current["general"], int(row["general"]))
            current["grocery"] = max(current["grocery"], int(row["grocery"]))
        return rollups

    def _load_target_zip_rollups(
        self,
        boundary_path: Path,
        jurisdiction_rates: dict[str, dict[str, Any]],
    ) -> dict[str, dict[str, int]]:
        zip_rates: dict[str, dict[str, int]] = {}
        with boundary_path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            header = next(reader)
            zip_idx = header.index("Zip Code")
            state_idx = header.index("FIPS State Code")
            for row in reader:
                zip_code = str(row[zip_idx] or "").strip()[:5]
                if zip_code not in self._TARGET_ZIPS:
                    continue
                state_fips = str(row[state_idx] or "").strip()
                if state_fips != str(self._STATE_FIPS):
                    continue
                jurisdiction_code = ""
                for value in row:
                    candidate = str(value or "").strip()
                    if candidate in jurisdiction_rates:
                        jurisdiction_code = candidate
                        break
                if not jurisdiction_code:
                    continue
                rates = jurisdiction_rates.get(jurisdiction_code)
                if rates is None:
                    continue
                current = zip_rates.setdefault(zip_code, {"general": 0, "grocery": 0})
                current["general"] = max(current["general"], int(rates["general"]))
                current["grocery"] = max(current["grocery"], int(rates["grocery"]))
        return zip_rates

    def _pct_decimal_to_bps(self, value: Any) -> int:
        return int(round(float(value or 0) * 10000))

    def _pct_string_to_bps(self, value: Any) -> int:
        text = str(value or "").strip().replace("%", "")
        if not text:
            return 0
        return int(round(float(text) * 100))

    def _to_int(self, value: Any) -> int:
        text = str(value or "").strip()
        if not text:
            return 0
        try:
            return int(float(text))
        except ValueError:
            return 0


def source_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()
